"""
数据导出管理器
支持导出为 Excel、CSV、PDF 格式
"""
import csv
import io
import logging
from datetime import datetime, timedelta
from typing import List, Dict, Any, Optional
from enum import Enum

# 尝试导入可选依赖
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import inch, cm
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Table, TableStyle, Spacer, Image
    from reportlab.graphics.shapes import Drawing
    from reportlab.graphics.charts.barcharts import VerticalBarChart
    from reportlab.graphics.charts.linecharts import HorizontalLineChart
    from reportlab.graphics.charts.piecharts import Pie
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False


class ExportFormat(Enum):
    """导出格式枚举"""
    CSV = "csv"
    EXCEL = "excel"
    PDF = "pdf"


class ExportTemplate(Enum):
    """PDF报表模板枚举"""
    STANDARD = "standard"
    VISUAL = "visual"
    FULL = "full"


class ExportManager:
    """数据导出管理器"""
    
    def __init__(self, data_manager):
        """
        初始化导出管理器
        
        Args:
            data_manager: 数据管理器实例
        """
        self.data_manager = data_manager
        self.logger = logging.getLogger(__name__)
    
    def export_to_csv(self, data: List[Dict], include_headers: bool = True) -> bytes:
        """
        导出数据为 CSV 格式
        
        Args:
            data: 要导出的数据列表
            include_headers: 是否包含表头
            
        Returns:
            CSV 字节数据
        """
        if not data:
            return b""
        
        output = io.StringIO()
        
        # 获取所有列名
        if data:
            fieldnames = list(data[0].keys())
        else:
            fieldnames = []
        
        writer = csv.DictWriter(output, fieldnames=fieldnames)
        
        if include_headers:
            writer.writeheader()
        
        for row in data:
            # 处理特殊字符和编码
            processed_row = {}
            for key, value in row.items():
                if isinstance(value, str):
                    # 替换可能导致CSV问题的字符
                    value = value.replace('\n', ' ').replace('\r', ' ')
                processed_row[key] = value
            writer.writerow(processed_row)
        
        # 返回UTF-8 with BOM编码（Excel兼容性）
        return output.getvalue().encode('utf-8-sig')
    
    def export_to_excel(self, data: List[Dict], sheet_name: str = "数据") -> bytes:
        """
        导出数据为 Excel 格式
        
        Args:
            data: 要导出的数据列表
            sheet_name: 工作表名称
            
        Returns:
            Excel 字节数据
        """
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl 库未安装，请运行: pip install openpyxl")
        
        if not data:
            return b""
        
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name[:31]
        
        # 定义样式
        header_font = Font(bold=True, size=11)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font_white = Font(bold=True, size=11, color="FFFFFF")
        center_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 获取列名
        if data:
            headers = list(data[0].keys())
        else:
            headers = []
        
        # 写入表头
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=str(header))
            cell.font = header_font_white
            cell.fill = header_fill
            cell.alignment = center_alignment
            cell.border = thin_border
        
        # 写入数据
        for row_idx, row_data in enumerate(data, 2):
            for col_idx, key in enumerate(headers, 1):
                value = row_data.get(key, "")
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                
                # 数字类型处理
                if isinstance(value, (int, float)):
                    cell.number_format = '#,##0' if isinstance(value, int) else '#,##0.00'
        
        # 自动调整列宽
        for col_idx, header in enumerate(headers, 1):
            column_letter = get_column_letter(col_idx)
            max_length = max(
                len(str(header)),
                max(len(str(row.get(header, ""))) for row in data) if data else 0
            )
            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
        
        # 冻结首行
        ws.freeze_panes = 'A2'
        
        # 保存到字节流
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output.getvalue()
    
    def export_to_pdf(self, 
                       data: List[Dict], 
                       template: ExportTemplate = ExportTemplate.STANDARD,
                       report_title: str = "数据导出报告",
                       stats_data: Optional[Dict] = None,
                       chart_data: Optional[Dict] = None) -> bytes:
        """
        导出数据为 PDF 格式
        
        Args:
            data: 要导出的数据列表
            template: 报表模板类型
            report_title: 报告标题
            stats_data: 统计数据（用于模板展示）
            chart_data: 图表数据（用于可视化模板）
            
        Returns:
            PDF 字节数据
        """
        if not REPORTLAB_AVAILABLE:
            raise ImportError("reportlab 库未安装，请运行: pip install reportlab")
        
        output = io.BytesIO()
        
        # 创建文档
        doc = SimpleDocTemplate(
            output,
            pagesize=A4,
            rightMargin=36,
            leftMargin=36,
            topMargin=36,
            bottomMargin=36
        )
        
        # 样式
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            spaceAfter=20,
            alignment=1
        )
        subtitle_style = ParagraphStyle(
            'Subtitle',
            parent=styles['Heading2'],
            fontSize=12,
            spaceAfter=10,
            textColor=colors.HexColor('#666666')
        )
        
        # 构建内容
        elements = []
        
        # 标题
        elements.append(Paragraph(report_title, title_style))
        elements.append(Paragraph(f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", subtitle_style))
        elements.append(Spacer(1, 20))
        
        # 根据模板添加内容
        if template in [ExportTemplate.STANDARD, ExportTemplate.FULL]:
            # 添加统计摘要
            if stats_data:
                elements.append(Paragraph("统计摘要", styles['Heading2']))
                elements.append(Spacer(1, 10))
                
                stats_table_data = [["指标", "数值"]]
                stats_table_data.extend([
                    ["总数据量", str(stats_data.get('total_count', 0))],
                    ["时间范围", stats_data.get('date_range', '全部')],
                    ["房间数量", str(stats_data.get('room_count', 1))]
                ])
                
                stats_table = Table(stats_table_data, colWidths=[2*inch, 2*inch])
                stats_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 12),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#F5F5F5')),
                    ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#CCCCCC')),
                ]))
                elements.append(stats_table)
                elements.append(Spacer(1, 20))
        
        # 添加图表（可视化和完整模板）
        if template in [ExportTemplate.VISUAL, ExportTemplate.FULL]:
            if chart_data:
                elements.append(Paragraph("数据可视化", styles['Heading2']))
                elements.append(Spacer(1, 10))
                
                # 柱状图
                if chart_data.get('bar_data'):
                    try:
                        drawing = Drawing(400, 200)
                        bc = VerticalBarChart()
                        bc.x = 50
                        bc.y = 50
                        bc.height = 125
                        bc.width = 300
                        
                        bar_data = chart_data['bar_data']
                        bc.data = [bar_data.get('values', [])]
                        bc.categoryAxis.categoryNames = bar_data.get('labels', [])
                        bc.bars[0].fillColor = colors.HexColor('#4472C4')
                        bc.valueAxis.valueMin = 0
                        
                        drawing.add(bc)
                        elements.append(drawing)
                        elements.append(Spacer(1, 20))
                    except Exception as e:
                        self.logger.warning(f"创建柱状图失败: {e}")
                
                # 饼图
                if chart_data.get('pie_data'):
                    try:
                        drawing = Drawing(400, 200)
                        pc = Pie()
                        pc.x = 150
                        pc.y = 25
                        pc.width = 100
                        pc.height = 100
                        
                        pie_data = chart_data['pie_data']
                        pc.data = pie_data.get('values', [])
                        pc.labels = pie_data.get('labels', [])
                        pc.slices.strokeWidth = 0.5
                        pc.slices[0].fillColor = colors.HexColor('#4472C4')
                        pc.slices[1].fillColor = colors.HexColor('#70AD47')
                        pc.slices[2].fillColor = colors.HexColor('#FFC000')
                        pc.slices[3].fillColor = colors.HexColor('#ED7D31')
                        pc.slices[4].fillColor = colors.HexColor('#5B9BD5')
                        
                        drawing.add(pc)
                        elements.append(drawing)
                        elements.append(Spacer(1, 20))
                    except Exception as e:
                        self.logger.warning(f"创建饼图失败: {e}")
        
        # 添加数据表格
        elements.append(Paragraph("详细数据", styles['Heading2']))
        elements.append(Spacer(1, 10))
        
        if data:
            # 准备表格数据
            headers = list(data[0].keys())
            table_data = [headers]
            
            # 限制数据量（PDF页面限制）
            for row in data[:500]:  # 最多显示500行
                table_data.append([str(row.get(h, "")) for h in headers])
            
            # 创建表格
            col_widths = [min(1.5 * inch, A4[0] / len(headers) - 10) for _ in headers]
            data_table = Table(table_data, colWidths=col_widths, repeatRows=1)
            
            # 表格样式
            style = [
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
                ('TOPPADDING', (0, 0), (-1, 0), 6),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CCCCCC')),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ]
            
            # 交替行颜色
            for i in range(1, len(table_data)):
                if i % 2 == 0:
                    style.append(('BACKGROUND', (0, i), (-1, i), colors.HexColor('#F5F5F5')))
            
            data_table.setStyle(TableStyle(style))
            elements.append(data_table)
            
            # 如果有更多数据，添加提示
            if len(data) > 500:
                elements.append(Spacer(1, 10))
                elements.append(Paragraph(
                    f"* 仅显示前500行，共 {len(data)} 行数据",
                    ParagraphStyle('Note', parent=styles['Normal'], textColor=colors.HexColor('#999999'))
                ))
        else:
            elements.append(Paragraph("暂无数据", styles['Normal']))
        
        # 添加页脚
        def add_footer(canvas, doc):
            canvas.saveState()
            canvas.setFont('Helvetica', 8)
            canvas.setFillColor(colors.HexColor('#999999'))
            page_num = canvas.getPageNumber()
            canvas.drawCentredString(A4[0]/2, 30, f"第 {page_num} 页")
            canvas.drawRightString(A4[0] - 36, 30, "鸽切监控系统")
            canvas.restoreState()
        
        doc.build(elements, onFirstPage=add_footer, onLaterPages=add_footer)
        output.seek(0)
        
        return output.getvalue()
    
    def get_export_data(self, 
                        rooms: List[int],
                        date_range: str = "all",
                        start_date: Optional[str] = None,
                        end_date: Optional[str] = None,
                        metrics: Optional[Dict[str, bool]] = None) -> List[Dict]:
        """
        获取要导出的数据
        
        Args:
            rooms: 房间ID列表
            date_range: 时间范围类型
            start_date: 开始日期（YYYY-MM-DD）
            end_date: 结束日期（YYYY-MM-DD）
            metrics: 要包含的指标
            
        Returns:
            格式化后的数据列表
        """
        if metrics is None:
            metrics = {'basic': True, 'rating': True, 'room': True}
        
        # 确定日期范围
        today = datetime.now().strftime('%Y-%m-%d')
        if date_range == "today":
            start_date = today
            end_date = today
        elif date_range == "7days":
            start_date = (datetime.now() - timedelta(days=7)).strftime('%Y-%m-%d')
            end_date = today
        elif date_range == "14days":
            start_date = (datetime.now() - timedelta(days=14)).strftime('%Y-%m-%d')
            end_date = today
        elif date_range == "30days":
            start_date = (datetime.now() - timedelta(days=30)).strftime('%Y-%m-%d')
            end_date = today
        
        # 获取数据
        all_data = []
        
        for room_id in rooms:
            # 从数据管理器获取数据
            if hasattr(self.data_manager, 'get_data_by_room_and_date'):
                if start_date and end_date:
                    # 如果有明确的日期范围
                    room_data = self.data_manager.get_data_by_room_and_date(room_id, start_date)
                    # 这里简化处理，实际应该处理日期范围
                else:
                    # 获取所有数据
                    room_data = self.data_manager.get_all_data()
                    # 按房间过滤
                    room_data = [item for item in room_data if item.get('room_id') == room_id]
            else:
                # 备用方式：获取所有数据
                room_data = self.data_manager.get_all_data()
                # 按房间过滤
                room_data = [item for item in room_data if item.get('room_id') == room_id]
            
            # 格式化数据
            for item in room_data:
                formatted = {}
                
                # 基础指标
                if metrics.get('basic'):
                    formatted.update({
                        '时间': item.get('time_display', item.get('timestamp', '-')),
                        '用户': item.get('username', '未知用户'),
                        '内容': item.get('content', '-')
                    })
                
                # 评分指标
                if metrics.get('rating'):
                    formatted['评分'] = item.get('rating', 0)
                
                # 房间信息
                if metrics.get('room'):
                    formatted['房间ID'] = room_id
                
                # 关键词匹配信息
                if 'keyword_matched' in item:
                    formatted['关键词匹配'] = '是' if item.get('keyword_matched') else '否'
                if 'matched_keyword' in item and item.get('matched_keyword'):
                    formatted['匹配关键词'] = item.get('matched_keyword', '-')
                
                all_data.append(formatted)
        
        return all_data
    
    def generate_filename(self, 
                          format_type: ExportFormat,
                          rooms: List[int],
                          include_timestamp: bool = True) -> str:
        """
        生成导出文件名
        
        Args:
            format_type: 格式类型
            rooms: 房间ID列表
            include_timestamp: 是否包含时间戳
            
        Returns:
            生成的文件名
        """
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S') if include_timestamp else ''
        
        if len(rooms) == 1:
            room_str = f"room_{rooms[0]}"
        else:
            room_str = f"rooms_{len(rooms)}"
        
        extensions = {
            ExportFormat.CSV: "csv",
            ExportFormat.EXCEL: "xlsx",
            ExportFormat.PDF: "pdf"
        }
        
        ext = extensions.get(format_type, "dat")
        
        if timestamp:
            return f"export_{room_str}_{timestamp}.{ext}"
        else:
            return f"export_{room_str}.{ext}"
